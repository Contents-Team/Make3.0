"""Build Make Naming Convention v8 spec + example guide as Excel workbook.

Output: doc/assets/MakeNamingConvention_v8.xlsx

Regenerable companion to doc/MAKE_FORMAT.md (v8). Non-developer guide workbook;
minor styling may differ from any hand-tuned copy.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "doc" / "assets" / "MakeNamingConvention_v8.xlsx"

# Styles
HDR_FONT = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(name="Malgun Gothic", size=14, bold=True, color="2F5496")
SECTION_FONT = Font(name="Malgun Gothic", size=12, bold=True, color="1F3864")
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")
BAD_FILL = PatternFill("solid", fgColor="FCE4D6")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def write_table(ws, start_row, headers, rows, col_widths=None, row_fills=None):
    """Write a table starting at start_row. Returns next free row."""
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=ci, value=h)
    style_header_row(ws, start_row, len(headers))
    for ri, row in enumerate(rows, start=start_row + 1):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
            cell.font = Font(name="Malgun Gothic", size=10)
            if row_fills and (ri - start_row - 1) < len(row_fills):
                fill = row_fills[ri - start_row - 1]
                if fill:
                    cell.fill = fill
    if col_widths:
        for ci, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[start_row].height = 22
    return start_row + len(rows) + 2


def title(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    return row + 2


def section(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = SECTION_FONT
    return row + 1


def note(ws, row, text, span=4):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Malgun Gothic", size=9, italic=True, color="595959")
    cell.alignment = WRAP
    cell.fill = NOTE_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 30
    return row + 2


wb = Workbook()

# ─────────────────────────────────────────────────────────────
# Sheet 1 — Overview
# ─────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Overview"
r = title(ws, 1, "Make Asset Naming Convention v8")
ws.cell(row=r, column=1, value="문서 버전: v8 / Repo: Make3.0 v0.1.7 / 작성일: 2026-07-01").font = Font(name="Malgun Gothic", size=10, italic=True)
r += 2

r = section(ws, r, "[ 이 문서를 읽기 전에 — 용어 안내 ]")
glossary = [
    ("원본 에셋 (Source Asset)", "모델러·디자이너가 만들어 전달하는 원본 파일. 예: .fbx, .png, .wav"),
    ("인스턴스 (Instance)", "Make Editor 안에 배치된 객체. 원본 파일을 씬에 가져다 놓은 것"),
    ("씬 트리 (Scene Tree)", "Make Editor 의 왼쪽 계층 구조. 부모-자식 관계로 객체가 묶여 있음"),
    ("앞 약자 / 뒤 약자", "이름 맨 앞 또는 맨 뒤에 붙는 짧은 표시 글자 (예: GNB_, _On)"),
    ("Vocab (정해진 단어 모음)", "허용된 단어 목록. 이 안에서만 골라 써야 함"),
    ("T1 / T2", "이 문서가 정한 두 단계: 원본 파일 / 배치된 인스턴스"),
]
r = write_table(ws, r, ["용어", "쉽게 풀어 쓰기"], glossary, col_widths=[28, 80])

r = section(ws, r, "[ 이 규칙이 따르는 5가지 큰 원칙 ]")
principles = [
    ("1. 어디서 보는지에 따라 다르게", "OS 파일 탐색기 / Make Editor / 코드 검색 — 각자 보기 좋은 모양이 다름"),
    ("2. Make Editor 안의 이름은 짧게", "Make Editor 가 이미 아이콘으로 종류를 알려주므로 이름에 종류 표시 약자는 생략"),
    ("3. 부모-자식 관계는 트리로 표현", "부모 이름을 자식 이름에 반복해 적지 않음 — 트리가 이미 알려주니까"),
    ("4. 구분 기호는 언더바 _ 하나만", "이중 언더바 __ 와 하이픈 - 은 모두 사용 금지"),
    ("5. 약자로 짧게 쓰기", "상태·동작·방향 모두 정해진 짧은 약자로 (3글자 또는 1글자)"),
]
r = write_table(ws, r, ["원칙", "쉬운 설명"], principles, col_widths=[34, 80])

r = section(ws, r, "[ 2단계 (T1 / T2) ]")
tiers = [
    ("T1", "원본 파일", "OS 파일 탐색기에서 보임", "PascalCase + 확장자 + (필요 시) 카테고리 약자", "GlassBottle_v01.fbx, Gnb_01_On.png, Bgm_Tutorial_01.wav"),
    ("T2", "Make Editor 인스턴스", "Make Editor 의 씬 트리에서 보임", "PascalCase, 종류 약자(mdl_ 등) 없음, 토큰은 언더바 _ 로 구분", "Btn_Start, Cupcake, Gnb_01_On"),
]
r = write_table(ws, r, ["단계", "무엇인지", "어디서 보이는지", "쓰는 모양", "예시"], tiers,
                col_widths=[6, 18, 34, 44, 50])

r = section(ws, r, "[ 시트 안내 ]")
sheets_info = [
    ("Overview", "이 시트 — 큰 원칙과 2단계 요약"),
    ("T1 Source", "원본 파일 이름 규칙 (3D 모델·이미지·소리·영상·폰트)"),
    ("T2 Instance", "Make Editor 인스턴스 이름 규칙 + 자주 쓰는 단어 모음"),
    ("Animation", "애니메이션 이름 규칙 + 동작·방향 약자 모음"),
    ("Vocab Tables", "약자·단어 모음 한눈에 보기"),
    ("Examples", "실제 파일을 v8 규칙으로 바꾼 비교표"),
    ("Checklist", "내보내기 전에 확인할 항목 (체크리스트)"),
    ("Anti-Patterns", "이렇게 쓰면 안 되는 예시"),
]
r = write_table(ws, r, ["시트", "내용"], sheets_info, col_widths=[18, 70])

r = section(ws, r, "[ C# 컨벤션 대조 (Microsoft .NET) ]")
r = note(ws, r, "에셋·씬·파일 이름은 C# 식별자가 아니므로 전면 적용은 부적절합니다. 그대로 전이되는 원칙(선명도·__ 금지·단일문자 지양·PascalCase)은 따르고, 안 맞는 항목(_ 구분자·패밀리 접두사)은 의도적 일탈로 이유를 명기합니다. 출처: Microsoft Learn — C# 식별자 명명 규칙 / .NET 코딩 규칙")
cs = [
    ("기본 표기", "PascalCase", "형식·public 멤버 PascalCase", "✅ 일치"),
    ("이중 밑줄 __", "자체 금지 (B19)", "__ 금지 (컴파일러 예약)", "✅ 준수 — 이중 밑줄 예시 1곳 수정 완료"),
    ("약어 Btn/Nav/Clr/Scl", "적극 사용", "약어 지양·\"간결성보다 선명도\"", "⚠️ 의도적 일탈 — 에디터 표시명 축약(원칙2), Vocab 용어집으로 완화"),
    ("단일문자 방향 L/R/U/D", "사용", "단일문자 금지 (루프 카운터만 예외)", "⚠️ 일탈 — 자산 suffix 관용이나 명시 필요"),
    ("_ 단어 구분자", "Btn_Navi_Left_Pressed", "_ 는 private 필드 접두사로만 사용", "⚠️ 자산·파일명엔 적합 / 코드 심볼은 무언더바 PascalCase"),
    ("약자 대소문자 (GNB/BGM/SFX)", "→ Gnb/Bgm/Sfx (PascalCase)", "3글자+ 약어도 PascalCase", "✅ 채택 — 3글자+ 약어 PascalCase(Gnb/Bgm/Info), 2글자(UI)는 대문자 유지"),
]
r = write_table(ws, r, ["항목", "현재 컨벤션", "MS C# 규칙", "판정 / 조치"], cs,
                col_widths=[24, 26, 32, 44])

r = section(ws, r, "[ 개선 제안 ]")
suggest = [
    ("1. 코드 경계 정의 — 카테고리(Bgm/Sfx/Info/Bgv)·애니메이션이 C# enum/직렬화 필드가 되는 지점엔 PascalCase 병행형 제공 (AudioCategory.Bgm)",),
    ("2. 각 의도적 일탈 항목에 \"이유\" 한 줄 부착 (예: _ 구분자는 에디터 가독성용)",),
    ("3. 약자 대소문자 규칙 채택: 3글자+ PascalCase(Gnb/Bgm), 2글자 대문자(UI) — Vocab·전 예시 반영 완료",),
]
for s in suggest:
    cell = ws.cell(row=r, column=1, value=s[0])
    cell.font = Font(name="Malgun Gothic", size=10)
    cell.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1

# ─────────────────────────────────────────────────────────────
# Sheet 2 — T1 Source Asset
# ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("T1 Source")
r = title(ws, 1, "T1 — 원본 파일 이름 규칙")
r = note(ws, r, "모델러·디자이너가 만들어서 전달하는 원본 파일(작업 폴더에 있는 파일)에 적용. PascalCase 로 쓰고 확장자(.fbx, .png 등)를 붙임. 확장자가 종류를 알려주므로 이름 앞에 종류 약자(mdl_ 등)는 생략. 단, PNG·WAV·MP4 처럼 같은 확장자가 여러 용도로 쓰이는 경우에만 카테고리 약자를 이름 안에 넣음.")

r = section(ws, r, "[ 공통 규칙 ]")
rules = [
    ("이름에 공백을 넣지 않는다", "GlassBottle_v01.fbx", "Glass Bottle_v01.fbx"),
    ("구분 기호는 언더바 _ 만 (하이픈 - 금지)", "Contents_3_1_1", "Contents_3-1-1-2"),
    ("이름은 PascalCase 로 붙임 (예: GlassBottle). 언더바 _ 는 구조 토큰(종류·이름·상태·번호·방향) 구분에만 사용 (예: GlassBottle_Open, Gnb_01)", "BatteryPack", "batterypack / battery_pack"),
    ("변형 번호는 두 자리로 통일", "_01, _02", "_1, _001"),
    ("버전은 v + 두 자리 (선택 사항)", "_v01, _v02", "_V1, _ver01"),
    ("확장자가 알려주는 종류를 이름에 또 적지 않는다", "GlassBottle_v01.fbx", "GlassBottle_Mod_v01.fbx"),
    ("영문으로만 적기 (한글 파일명 금지)", "BatteryPack_v01", "배터리팩_v01"),
]
r = write_table(ws, r, ["규칙", "✅ 이렇게 쓰세요", "❌ 이렇게 쓰지 마세요"], rules,
                col_widths=[54, 36, 36])

r = section(ws, r, "[ 파일 종류별 규칙 ]")

r = section(ws, r, "■ 3D 모델 (.fbx)")
r = note(ws, r, "확장자 .fbx 가 \"3D 모델\"이라는 종류를 알려줍니다. 이름에 또 적지 마세요. 자식 부품을 별도 파일로 보낼 때는 부모 이름을 함께 적어 어떤 모델의 부품인지 알 수 있게 하세요. 애니메이션이 포함된 모델은 이름 앞에 접두사 Ani_ 를 붙여 애니메이션 없는 모델과 구분합니다 (예: Ani_Cupcake.fbx). 애니메이션이 없는 모델은 접두사 없이 그대로 둡니다.")
fbx = [
    ("GlassBottle_v01.fbx", "기본 모델"),
    ("GlassBottle_v02.fbx", "수정된 버전 (v02)"),
    ("GlassBottle_Open_v01.fbx", "다른 형태 (뚜껑 열린 버전)"),
    ("GlassBottle_Cap_v01.fbx", "자식 부품 — 부모 이름 포함"),
    ("BatteryPack_TopCover_v01.fbx", "자식 부품"),
    ("Ani_Cupcake_v01.fbx", "애니메이션 포함 모델 — 접두사 Ani_"),
]
r = write_table(ws, r, ["파일 이름", "쓰임새"], fbx, col_widths=[40, 50])

r = section(ws, r, "■ 이미지 — UI 스프라이트 (.png/.jpg)")
sprite = [
    ("Family + 이름 + 상태 (토큰마다 언더바 _ 로 구분)", "Gnb_01_On.png, Btn_Navi_Left_Pressed.png"),
    ("Family 만으로 충분할 때는 단독 사용 OK", "Logo.png, Tip.png, Pnl_Tablet.png"),
    ("상태는 정해진 목록 안에서만 사용 (기본 Default 는 생략, 다른 상태만 표시)", "Gnb_On.png, Gnb_Off.png"),
    ("방향은 풀어쓰기 (Left / Right / Up / Down)", "Btn_Navi_Left.png, Rotation_Up.png"),
]
r = write_table(ws, r, ["규칙", "예시"], sprite, col_widths=[54, 48])

r = section(ws, r, "■ 소리 (.wav / .mp3)")
r = note(ws, r, "형식: <카테고리>_<설명>_<번호>.확장자. 카테고리 약자와 설명 사이에 언더바 _ 를 한 번 넣습니다.")
audio = [
    ("Bgm", "배경 음악", "Bgm_Tutorial_01.wav, Bgm_Boss_01.wav"),
    ("Sfx", "효과음 (버튼 클릭·완료음·경고음 등)", "Sfx_UI_Click_01.wav, Sfx_Complete_01.wav"),
    ("Nar", "내레이션·더빙 (이전 이름: VO)", "Nar_Intro_01.wav, Nar_Step_01.wav"),
]
r = write_table(ws, r, ["카테고리", "의미", "예시"], audio, col_widths=[14, 44, 50])

r = section(ws, r, "■ 영상 (.mp4)")
video = [
    ("Info", "설명·정보 전달용 (컷씬·튜토리얼·인트로·아웃트로 통합)", "Info_Assembly_01.mp4"),
    ("Bgv", "배경용 (반복 재생되는 배경 영상)", "Bgv_Workshop.mp4"),
]
r = write_table(ws, r, ["카테고리", "의미", "예시"], video, col_widths=[14, 48, 46])

# ─────────────────────────────────────────────────────────────
# Sheet 3 — T2 Instance
# ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("T2 Instance")
r = title(ws, 1, "T2 — Make Editor 인스턴스 이름 규칙")
r = note(ws, r, "Make Editor 의 씬 트리에 보이는, 사용자가 직접 이름을 짓는 객체입니다. 트리 옆 아이콘이 종류를 알려주므로 이름에 종류 약자(mdl_, img_ 같은 것) 는 생략합니다. 부모-자식 관계는 트리가 이미 보여주므로 자식 이름은 짧게 적습니다.")

r = section(ws, r, "[ 가장 중요한 규칙 ]")
core = [
    ("단어마다 첫 글자 대문자 (PascalCase)", "Btn_Start (기본), Btn_Start_Pressed (눌림)", "btn_start, gnb_act"),
    ("타입 약자(mdl_, img_) 생략 (T2 에서만)", "Cupcake", "mdl_cupcake"),
    ("토큰 사이를 언더바 _ 하나로 구분 (Family 약자 뒤에도)", "Contents_3_1_1", "Contents__3, Contents-3"),
    ("이름 끝에 확장자 안 붙이기", "Logo_x4", "Logo_x4.png"),
    ("정해진 4방향 자식은 방향 단어로 (Left/Right/Up/Down)", "RotationPanel 아래에 Left/Right/Up/Down", "(자유롭게 이름 짓는 형제 노드 사이에선 금지)"),
    ("자유 명명일 때는 의미 있는 이름", "Btn_Reset, Btn_ZoomIn", "R, Z (무슨 뜻인지 모름)"),
    ("상태는 기본(Default) 생략 — 변형 상태만 뒤에 표시", "Btn_Start → Btn_Start_Pressed", "Btn_Start_Default (기본인데 표시)"),
    ("장면 구분: 최상위 컨테이너 번호 (World_01 / Screen_01). 깊은 콘텐츠 경로는 숫자 최대 3자리", "World_01, Screen_01 · Contents_3_1_1", "Contents_3_1_1_2 (콘텐츠 4자리 초과)"),
    ("애니메이션 포함 3D 모델은 접두사 Ani_", "Ani_Cupcake", "Cupcake (애니메이션인데 구분 표시 없음)"),
]
r = write_table(ws, r, ["규칙", "✅ 이렇게 쓰세요", "❌ 이렇게 쓰지 마세요"], core, col_widths=[44, 40, 40])

r = section(ws, r, "[ 이름 짜는 순서 ]")
ws.cell(row=r, column=1, value="<Family>_<고유 이름>[_<방향>][_<상태>]").font = Font(name="Consolas", size=11, bold=True)
r += 2
syntax = [
    ("Family 약자", "어떤 UI 인지 알려주는 짧은 약자 (선택 사항, 권장)", "Gnb, Snb, Bnb, Btn, Pnl, ..."),
    ("고유 이름", "이 객체만의 이름 (PascalCase)", "Start, Navi, Rotation"),
    ("방향", "방향 정보가 있을 때 풀어서 (Left/Right/Up/Down)", "Left, Right, Up, Down, Clockwise, AntiClockwise"),
    ("상태", "상태 단어 (기본 Default 는 생략)", "Pressed, On, Off, Disabled (Default 는 안 붙임)"),
]
r = write_table(ws, r, ["부분", "설명", "예시"], syntax, col_widths=[18, 44, 40])

r = section(ws, r, "[ 컴포넌트 서브오브젝트 (v8) — 버튼/메뉴 한 덩어리 구성 ]")
r = note(ws, r, "버튼·메뉴처럼 이미지+글자+인터랙션이 한 덩어리인 UI 는 메뉴별 '그룹(빈 오브젝트)' 아래에 역할별로 나눠 넣습니다. 자식 이름은 부모 이름을 물려받아 <부모>_<역할> 로 짓습니다 (예: Gnb_Home_Text). 이렇게 하면 Gnb_Home 하나로 관련 객체가 다 검색되고, 사용자도 어디 소속인지 바로 압니다.", span=3)
roles = [
    ("_Img", "버튼 이미지 (모양 무관, 상태별 그림 교체)", "Gnb_Home_Img"),
    ("_Text", "글자 라벨 (예전 _Label)", "Gnb_Home_Text"),
    ("_Ico", "아이콘", "Gnb_Settings_Ico"),
    ("_Btn", "인터랙션(클릭·이벤트) 투명 사각형", "Gnb_Home_Btn"),
    ("_Bg", "실제 전체 배경일 때만", "Screen_Bg"),
]
r = write_table(ws, r, ["역할", "의미", "예시 (부모=Gnb_Home)"], roles, col_widths=[12, 46, 30])
r = note(ws, r, "이벤트 붙이는 곳: 클릭·게이즈 같은 인터랙션은 '언제나' 투명 사각형 *_Btn 에 붙입니다 (이벤트 찾기 = *_Btn 찾기). *_Btn 은 그림보다 살짝 크게(여유 마진) 만들어 주변 영역까지 눌리게 합니다 (VR 손·시선 오차 흡수).", span=3)

r = section(ws, r, "[ 겹침(잘림·z-fighting) 방지 — Z 위치 (v8) ]")
zlayer = [
    ("_Btn (투명, 맨 앞)", "-0.02", "입력을 먼저 가로챔. 투명이라 글자 안 가림"),
    ("_Text / _Ico (중간)", "-0.01", "그림보다 앞"),
    ("_Img (기준면, 맨 뒤)", "0", "기준"),
]
r = write_table(ws, r, ["레이어", "로컬 Z", "설명"], zlayer, col_widths=[26, 12, 46])
r = note(ws, r, "값(Δz)은 프로젝트 크기에 맞춰 0.001~0.05 사이에서 하나로 통일해 씁니다. '앞(카메라 쪽)=Z 작아짐' 부호도 프로젝트 전체가 동일하게 씁니다. 그룹(빈 오브젝트)은 위치를 갖고 자식은 그 안에서 상대 위치만 가지므로, 메뉴를 옮기면 그림·글자·버튼이 함께 따라갑니다.", span=3)

r = section(ws, r, "[ 이름이 길 때 줄이는 법 (v8) ]")
shorten = [
    ("부모에 이미 있는 말은 자식에 반복 안 함", "Gnb_Settings", "Gnb_SettingsMenu"),
    ("핵심 명사 하나로", "Gnb_Profile", "Gnb_UserProfileSettings"),
    ("정해진 약어만 사용", "Config, Info, Alert", "임의 약어"),
    ("최대 2단어·약 12자", "UserGuide", "너무 긴 이름"),
    ("긴 정식 명칭은 이름 말고 표시명(DisplayTitle)에", "이름=Gnb_Notice / 표시명=공지사항 알림 센터", ""),
]
r = write_table(ws, r, ["규칙", "✅ 이렇게", "❌ 이렇게 말고"], shorten, col_widths=[42, 40, 30])

r = section(ws, r, "[ 한글은 어디에? (v8) ]")
r = note(ws, r, "오브젝트 '이름(name)'은 영문 PascalCase 로 고정하고, 화면에 보일 한글은 '표시명(DisplayTitle)' 칸에 따로 적습니다. 이름을 한글로 쓰면 글자 깨짐(맥/윈도우 차이)·검색/코드 문제·AI 번역 부담이 생깁니다. Make 3.0 에 표시명 칸이 있으면 이름(영문)+표시명(한글) 둘 다 적고, 없으면 종류 약자(Gnb/_Img/_Btn 등)는 영문 그대로 두고 메뉴명 부분만 한글로 씁니다 (예: Gnb_홈_Btn). 나중에 도구가 영문으로 정리합니다.", span=3)

# ─────────────────────────────────────────────────────────────
# Sheet 4 — Hierarchy (씬 하이어아키 구조)
# ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Hierarchy")
r = title(ws, 1, "씬 하이어아키 구조 (초안 — 검토용)")
ws.cell(row=r, column=1, value="v8 · 2026-07-01 · Unity 씬 구조 모범사례 기반. 분류·이름 확인 후 수정 요청 주세요.").font = Font(name="Malgun Gothic", size=10, italic=True)
r += 1
r = note(ws, r, "목적: 빈 게임오브젝트(Empty)를 \"폴더 컨테이너\"로 써서 오브젝트의 형태(증강=월드공간 / 스크린=UI공간)별로 묶어 둡니다. ① AI·사람이 구조만 보고 빠르게 검색 ② 그룹 단위 접기/숨기기로 간편한 유지보수 가 목표입니다. ※ 이 트리는 표준 템플릿이며, 장면·프로젝트에서 쓰지 않는 컨테이너는 삭제하고 씁니다 (World·Screen 모두 접두사 없는 타입 컨테이너 방식 동일).")

r = section(ws, r, "[ 씬 트리 — 디바이스별 ]")
tree = [
    ("■ HMD — World 전용 (모델·UI 모두 월드 공간)", "", "", ""),
    ("World_01", "장면 단위 (World_NN)", "", ""),
    ("  ├─ Models", "3D 모델 (Ani_ 포함/미포함)", "Ani_Cupcake", "삭제 가능"),
    ("  ├─ Gnb", "전역 내비", "Gnb_01_On", "삭제 가능"),
    ("  ├─ Snb", "측면 내비", "Snb_01_On", "삭제 가능"),
    ("  ├─ Bnb", "하단 내비", "Bnb_01_On", "삭제 가능"),
    ("  ├─ Modal", "모달·팝업", "Modal_Exit", "삭제 가능"),
    ("  ├─ Btn", "일반 버튼", "Btn_Start", "삭제 가능"),
    ("  └─ Audio", "사운드 (Bgm·Sfx·Nar)", "Bgm_Workshop", "삭제 가능"),
    ("■ PC · Tablet — Models만 World, 나머지 Screen", "", "", ""),
    ("World_01", "3D 모델만 월드 공간", "", ""),
    ("  └─ Models", "3D 모델 (Ani_ 포함/미포함)", "Ani_Cupcake", "삭제 가능"),
    ("Screen_01", "UI·사운드 (스크린 공간)", "", ""),
    ("  ├─ Gnb", "전역 내비", "Gnb_01_On", "삭제 가능"),
    ("  ├─ Snb", "측면 내비", "Snb_01_On", "삭제 가능"),
    ("  ├─ Bnb", "하단 내비", "Bnb_01_On", "삭제 가능"),
    ("  ├─ Modal", "모달·팝업", "Modal_Exit", "삭제 가능"),
    ("  ├─ Btn", "일반 버튼", "Btn_Start", "삭제 가능"),
    ("  └─ Audio", "UI 사운드 · Nar(TTS 시 Screen 텍스트)", "Sfx_Click_01, Nar_Intro_01", "삭제 가능"),
]
r = write_table(ws, r, ["구조", "설명", "예시", "미사용 시"], tree,
                col_widths=[42, 32, 26, 14])

r = section(ws, r, "[ 컨테이너 규칙 ]")
container = [
    ("최상위 컨테이너는 빈 오브젝트, Transform 원점(0,0,0)·Scale 1 고정", "자식 오브젝트의 월드 좌표가 틀어지지 않도록"),
    ("컨테이너 안 분류는 접두사 없이 패밀리/타입 이름 (Models, Gnb, Audio) · PascalCase", "AI·검색이 형태와 종류를 바로 인식"),
    ("그룹 안에 객체를 자식(child)으로 넣어 정리 → 접기/숨기기로 관리", "씬이 커져도 한눈에 보고 빠르게 찾음"),
    ("순수 정리용 빈 오브젝트에는 EditorOnly 태그 → 빌드에서 자동 제외", "런타임 메모리·계층 영향 최소화"),
    ("빈 부모 만들기: 선택 후 Ctrl+Shift+G (Mac: Cmd+Shift+G)", "Unity 기본 단축키"),
    ("장면은 최상위 컨테이너 뒤 번호로 구분 (World_01 / Screen_01) — 같은 번호 = 한 장면", "이름만 보고 장면 파악 · 월드·스크린을 같은 번호로 묶음 (Audio는 각 안에 포함)"),
]
r = write_table(ws, r, ["규칙", "이유"], container, col_widths=[56, 50])

r = section(ws, r, "[ 메뉴 묶기 · 이동 단위 (v8) ]")
menugrp = [
    ("메뉴 버튼은 번호(Gnb_01) 대신 이름 그룹(Gnb_Home)으로 묶는다", "무슨 메뉴인지 이름만 보고 알고, 그룹 하나로 관련 객체가 다 검색됨"),
    ("메뉴 버튼은 자기 패밀리(Gnb) 안에 둔다 — 버튼만 따로 모으지 않음", "위치·글자가 바뀌어도 이미지·글자·이벤트가 함께 움직이고 함께 찾아짐"),
    ("그룹(빈 오브젝트)이 위치를 갖고, 자식은 그 안에서 상대 위치", "메뉴 하나를 옮기면 하위 전체가 따라옴 (연결 유지)"),
    ("최상위 컨테이너(Screen_01/World_01/Gnb)만 원점 고정, 메뉴 그룹은 이동 단위", "기준 좌표는 최상위에서만 고정"),
]
r = write_table(ws, r, ["규칙", "이유"], menugrp, col_widths=[56, 50])

r = section(ws, r, "[ 정렬·간격 (v8) ]")
r = note(ws, r, "같은 바(Gnb/Snb/Bnb) 안의 메뉴들은 상위 컨테이너가 정렬·간격 규칙을 가지고 자동 배치합니다. 씬 트리에서의 순서 = 화면에서의 순서. 버튼을 넣고 빼면 자동으로 다시 정렬됩니다.", span=2)
layout = [
    ("Direction (방향)", "상단 Gnb·하단 Bnb → 가로 / 측면 Snb·Lnb → 세로", "바 종류로 결정"),
    ("Align (정렬)", "가로바: 왼쪽/가운데/오른쪽 · 세로바: 위/가운데/아래", "기본 가운데(Center)"),
    ("Spacing (간격)", "메뉴 사이 균등 간격 (고정값)", "프로젝트 전역 통일값"),
]
r = write_table(ws, r, ["속성", "값", "기본"], layout, col_widths=[20, 52, 28])

r = section(ws, r, "[ Unity 참고 ]")
unity_notes = [
    "• 빈 게임오브젝트를 폴더처럼 써서 객체를 그룹화하면 접기/숨기기가 쉬움",
    "• 단순 구분선(헤더) 빈 객체 아래 오브젝트는 자식이 아님 — 자식 중첩 시 좌표 틀어짐 주의",
    "• 본 초안은 '자식 컨테이너' 방식(원점 고정)을 채택 — 그룹 검색·유지보수에 유리",
    "출처: Unity Manual(Hierarchy), Unity Learn(Organizing your Scene), gamedevbeginner.com",
]
for n in unity_notes:
    cell = ws.cell(row=r, column=1, value=n)
    cell.font = Font(name="Malgun Gothic", size=10)
    cell.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1

# ─────────────────────────────────────────────────────────────
# Sheet 5 — Animation
# ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Animation")
r = title(ws, 1, "애니메이션 이름 규칙 (v8)")
r = note(ws, r, "애니메이션은 '애니메이션 목록'이라는 별도 공간에 있어 종류 약자가 필요 없습니다. 이름은 대상 오브젝트와 무관하게 '무엇이 어떻게 움직이는지'를 의미 단어로 짓습니다 (가독성 우선, 약어 사용 안 함). 하나의 애니메이션이 여러 속성을 동시에 바꿀 수 있습니다.")

r = section(ws, r, "[ 움직일 수 있는 속성 (4가지 · 기술 기준) ]")
props = [
    ("Position", "위치 이동", "가능"),
    ("Rotation", "회전", "가능"),
    ("Scale", "크기 변화", "가능"),
    ("Color", "색상 변화", "불가 (UI·2D 전용)"),
]
r = write_table(ws, r, ["종류", "의미", "3D 모델 적용"], props, col_widths=[16, 28, 24])

r = section(ws, r, "[ 방향 (Position·Rotation 에 사용) ]")
adir = [
    ("Left / Right / Up / Down", "좌 / 우 / 상 / 하"),
    ("Clockwise / AntiClockwise", "시계 / 반시계 방향 회전"),
]
r = write_table(ws, r, ["방향", "의미"], adir, col_widths=[30, 40])

r = section(ws, r, "[ 이름 형식 ]")
ws.cell(row=r, column=1, value="<효과 이름>[_<방향>]").font = Font(name="Consolas", size=11, bold=True)
r += 1
ws.cell(row=r, column=1, value="방향이 필요하면 뒤에 _Left / _Right / _Up / _Down (예: Spin_Left, Slide_Up)").font = Font(name="Malgun Gothic", size=10, italic=True)
r += 1
ws.cell(row=r, column=1, value="속성을 나열하지 않고, Bounce·ZoomIn·Pop 처럼 '어떤 애니메이션인지'가 바로 보이는 직관적 효과 이름을 씁니다. 대상 오브젝트 이름은 넣지 않습니다.").font = Font(name="Malgun Gothic", size=10, italic=True)
r += 2

r = section(ws, r, "[ 자주 쓰는 효과 이름 ]")
effects = [
    ("Bounce", "통통 튀어오름", "Position+Scale · 3D 가능"),
    ("ZoomIn / ZoomOut", "확대 / 축소", "Scale · 3D 가능"),
    ("Pop", "팝 하고 나타남", "Scale · 3D 가능"),
    ("Pulse", "커졌다 작아지기 반복", "Scale · 3D 가능"),
    ("Spin", "빙글 회전 (방향 가능)", "Rotation · 3D 가능"),
    ("Slide", "미끄러져 이동 (방향 가능)", "Position · 3D 가능"),
    ("Flip", "뒤집기 (방향 가능)", "Rotation · 3D 가능"),
    ("Fade", "서서히 나타남 / 사라짐", "Color/투명도 · 3D 불가 (UI·2D)"),
    ("Highlight", "색으로 강조", "Color · 3D 불가 (UI·2D)"),
]
r = write_table(ws, r, ["효과 이름", "느낌 / 동작", "사용 속성 · 3D"], effects, col_widths=[20, 28, 36])

# ─────────────────────────────────────────────────────────────
# Sheet 6 — Vocab Tables
# ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Vocab Tables")
r = title(ws, 1, "약자·단어 모음 한눈에 보기")

r = section(ws, r, "[ UI 종류 약자 (T2 인스턴스용, 권장) ]")
family = [
    ("Gnb", "Global Navigation Bar (전체 메뉴바)", "메뉴명 그룹 우선 (Gnb_Home), 번호는 대안 (Gnb_01)"),
    ("Snb", "Side Navigation Bar (측면 메뉴바)", "메뉴명 그룹 우선 (Snb_Layers)"),
    ("Bnb", "Bottom Navigation Bar (하단 메뉴바)", "메뉴명 그룹 우선 (Bnb_Play)"),
    ("Btn", "Button (버튼)", "이름 반드시 함께"),
    ("Modal", "Modal / Popup (모달·팝업 창)", "이름 또는 용도 (Modal_Exit)"),
    ("Nav", "Navigation control (앞/뒤 이동 버튼)", "이름 추가 안 해도 됨"),
    ("Pnl", "Panel / Background container (패널·배경판)", "이름 추가 안 해도 됨"),
    ("Ico", "Icon (아이콘)", "이름 반드시 함께"),
    ("Bg", "Background image (배경 이미지)", "이름 추가 안 해도 됨"),
    ("Logo", "Logo / Brand mark (로고)", "이름 추가 안 해도 됨"),
    ("Tip", "Tooltip (말풍선·도움말)", "이름 추가 안 해도 됨"),
    ("TTS_", "음성 변환(TTS) 적용 텍스트", "TTS 적용 텍스트만 접두사 TTS_ (예: TTS_StartHeadline). 일반·버튼 텍스트는 접두사 없이 직관적 PascalCase 이름 (예: Start)"),
]
r = write_table(ws, r, ["약자", "의미", "이름이 함께 필요한지"], family, col_widths=[10, 44, 48])

r = section(ws, r, "[ 컴포넌트 역할 약자 (v8) ]")
role_v = [
    ("_Img", "버튼 이미지 (모양 무관, 상태별 그림 교체)"),
    ("_Text", "글자 라벨 (예전 _Label)"),
    ("_Ico", "아이콘"),
    ("_Btn", "인터랙션(클릭·이벤트) 투명 사각형 — 이벤트는 항상 여기, 최전방 Z"),
    ("_Bg", "실제 전체 배경일 때만"),
]
r = write_table(ws, r, ["역할", "의미"], role_v, col_widths=[12, 70])

r = section(ws, r, "[ 상태 (풀어쓰기) ]")
state = [
    ("—", "Default", "기본 상태 (접미사 생략)"),
    ("_Pressed", "Pressed", "눌린 상태"),
    ("_On", "On", "켜짐 / 활성 (토글)"),
    ("_Off", "Off", "꺼짐 / 비활성 (토글 · 상호작용 가능)"),
    ("_Disabled", "Disabled", "사용 불가 (상호작용 막힘 · 흐리게)"),
]
r = write_table(ws, r, ["접미사", "상태", "의미"], state, col_widths=[14, 18, 40])
r = note(ws, r, "변경 이력: _Hover / _Focused / _Highlight 는 사용 안 함 (VR·터치 환경). _Active 는 _On 으로 통합.", span=3)

r = section(ws, r, "[ 방향 (풀어쓰기) ]")
dirs_full = [
    ("_Left", "Left (왼쪽)", "회전 · 이동"),
    ("_Right", "Right (오른쪽)", "회전 · 이동"),
    ("_Up", "Up (위)", "회전 · 이동"),
    ("_Down", "Down (아래)", "회전 · 이동"),
    ("_Clockwise", "Clockwise (시계 방향)", "회전"),
    ("_AntiClockwise", "Anti-clockwise (반시계 방향)", "회전"),
]
r = write_table(ws, r, ["접미사", "의미", "쓰는 곳"], dirs_full, col_widths=[16, 30, 24])

r = section(ws, r, "[ 오디오 카테고리 ]")
audio_c = [
    ("Bgm", "Background Music", "배경음악"),
    ("Sfx", "Sound Effect", "효과음 (UI 클릭·완료음·경고음 등)"),
    ("Nar", "Narration", "내레이션·더빙 (이전 이름: VO)"),
]
r = write_table(ws, r, ["카테고리", "풀이", "의미"], audio_c, col_widths=[12, 24, 44])

r = section(ws, r, "[ 비디오 카테고리 ]")
video_c = [
    ("Info", "설명·정보", "컷씬·튜토리얼·인트로·아웃트로 통합 (정보 전달 영상)"),
    ("Bgv", "배경 영상 (Background Video)", "반복 재생 배경 영상"),
]
r = write_table(ws, r, ["카테고리", "풀이", "의미"], video_c, col_widths=[12, 28, 44])

# ─────────────────────────────────────────────────────────────
# Sheet 7 — Examples
# ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Examples")
r = title(ws, 1, "실제 파일을 v8 규칙으로 바꾼 비교표")
r = note(ws, r, "실제 파일(Make Templete_20260512.make, 노드 513개)의 이름을 v8 규칙으로 바꾼 예시입니다. 각 행은 \"현재 이름 → v8 이름 + 왜 바꾸는지\" 입니다.", span=5)

r = section(ws, r, "[ 씬·페이지·계층 ]")
scene_ex = [
    ("SceneRoot", "SceneRoot", "T2 인스턴스", "시스템 예약 이름 — 그대로 유지"),
    ("Start", "Start", "T2 인스턴스", "이미 짧고 의미가 명확함 — 그대로 유지"),
    ("Contents 1-1", "Contents_1_1", "T2 인스턴스", "공백·하이픈 → 언더바로 통일"),
    ("Contents 3-1-1-2", "Contents_3_1_1", "T2 인스턴스", "4단계 위계도 언더바로 평탄화"),
    ("3D Model", "Stage_3D", "T2 인스턴스", "공백 제거 + 의미 있는 이름 부여"),
    ("0016", "Cupcake", "T2 인스턴스 (용어집 등록)", "숫자로 시작하는 이름 → 의미 있는 이름으로"),
]
r = write_table(ws, r, ["현재 이름", "v8 이름", "단계", "왜 바꾸나"], scene_ex,
                col_widths=[26, 28, 22, 36])

r = section(ws, r, "[ UI 이미지 인스턴스 ]")
ui_ex = [
    ("GNB_Active.png", "Gnb_Home_Img (On 스프라이트)", "T2", "메뉴 그룹 Gnb_Home 안 이미지, 상태는 스프라이트 교체"),
    ("GNB_default.png", "Gnb_Home_Img", "T2", "기본 상태 스프라이트 (그룹=Gnb_Home)"),
    ("GNB_Disabled.png", "Gnb_Home_Img (Disabled)", "T2", "비활성 스프라이트"),
    ("GNB_Pressed.png", "Gnb_Home_Btn (이벤트 면)", "T2", "인터랙션은 투명 _Btn 에"),
    ("LNB_Active.png", "Snb_On", "T2", ""),
    ("btn_start.png", "Btn_Start", "T2", "Family 약자 Btn 사용"),
    ("btn_rotation_off.png", "Btn_Rotation_Off", "T2", ""),
    ("btn_rotation_on.png", "Btn_Rotation_On", "T2", ""),
    ("Tablet.png", "Pnl_Tablet", "T2", "Family 약자 Pnl(패널) 사용"),
    ("Navi.png", "Pnl_Nav", "T2", "Navi → Nav 로 표기 통일"),
    ("Logox4.png", "Logo_x4", "T2", "Family 약자 Logo, 변형 표시는 _x4"),
    ("tooltip.png", "Tip", "T2", "Family 약자 Tip 단독 사용"),
    ("Tablet.png 3", "Pnl_Tablet_03", "T2", "Unity 가 자동으로 붙인 번호를 정리"),
]
r = write_table(ws, r, ["현재 이름", "v8 이름", "단계", "왜 바꾸나"], ui_ex,
                col_widths=[28, 24, 8, 50])

r = section(ws, r, "[ 방향이 있는 컨트롤 (형제 노드 한 글자 사용) ]")
dir_ex = [
    ("L_btn_navi", "Btn_Navi_Left", "T2", "방향을 앞이 아닌 뒤에 풀어서 (Left/Right/Up/Down)"),
    ("R_btn_navi", "Btn_Navi_Right", "T2", ""),
    ("L_btn_navi_Default.png", "Btn_Navi_Left", "T2", "방향 + 기본 상태(생략) → 방향만"),
    ("R_btn_navi_Presssed.png", "Btn_Navi_Right_Pressed", "T2", "오타 교정(Presssed→Pressed) + 풀어쓰기"),
    ("L_btn_navi_Highlight.png", "Btn_Navi_Left_Pressed", "T2", "Highlight 미사용 → Pressed 로 대체"),
    ("L_rotation", "Left", "T2 (형제 노드)", "부모(RotationPanel) 컨텍스트 + 방향 풀어쓰기"),
    ("R_rotation", "Right", "T2 (형제 노드)", ""),
    ("U_rotation", "Up", "T2 (형제 노드)", ""),
    ("D_rotation", "Down", "T2 (형제 노드)", ""),
    ("R_rotation_Active.png", "Right_On", "T2 (형제 노드)", "방향 + 상태"),
]
r = write_table(ws, r, ["현재 이름", "v8 이름", "단계", "왜 바꾸나"], dir_ex,
                col_widths=[30, 22, 18, 46])

r = section(ws, r, "[ 텍스트·말풍선 인스턴스 ]")
text_ex = [
    ("Text 1", "StartHeadline", "T2", "의미 있는 이름 부여 (T2 에선 txt_ 같은 종류 약자 안 씀). 음성 변환(TTS) 적용 텍스트는 접두사 TTS_ 사용 (예: TTS_StartHeadline)"),
    ("txt_Navi_Title", "NavTitle", "T2", "T2 에서는 txt_ 같은 종류 약자 생략"),
    ("txt_Navi_page", "NavPage", "T2", ""),
]
r = write_table(ws, r, ["현재 이름", "v8 이름", "단계", "왜 바꾸나"], text_ex,
                col_widths=[26, 22, 8, 58])

r = section(ws, r, "[ 애니메이션 (애니메이션 목록에 들어가므로 종류 약자 없음 · 효과 기반) ]")
anim_full = [
    ("new Clip (btn_start, 크기)", "Pop", "(애니메이션)", "효과 이름 (대상 이름 제외)"),
    ("new Clip (3D Model, 회전)", "Spin_Left", "(애니메이션)", "효과 + 방향 (대상 이름 제외)"),
    ("new Clip1 (3D Model, 회전)", "Spin_Right", "(애니메이션)", ""),
    ("new Clip2 (3D Model, 회전)", "Spin_Up", "(애니메이션)", ""),
    ("new Clip3 (3D Model, 회전)", "Spin_Down", "(애니메이션)", ""),
    ("new Clip4 (3D Model, 크기+이동)", "ZoomIn", "(애니메이션)", "확대 (이동+크기)"),
    ("new Clip5 (3D Model, 크기)", "ZoomOut", "(애니메이션)", "축소"),
]
r = write_table(ws, r, ["현재 이름", "v8 이름 (효과 기반)", "단계", "왜 바꾸나"], anim_full,
                col_widths=[36, 22, 18, 30])

# ─────────────────────────────────────────────────────────────
# Sheet 8 — Checklist
# ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Checklist")
r = title(ws, 1, "내보내기 전에 확인할 항목 (v8 체크리스트)")
r = note(ws, r, ".make 파일을 만들기 전에 한 번씩 확인하세요. 모든 항목에 ✓ 가 들어가면 v8 규칙을 잘 따른 파일입니다.", span=2)

r = section(ws, r, "[ T1 — 원본 파일 ]")
t1_check = [
    ("파일 이름에 공백이 없는가", ""),
    ("구분자가 언더바 _ 만 쓰이는가 (하이픈 - 없음)", ""),
    ("단어마다 첫 글자가 대문자인가 (PascalCase)", ""),
    ("변형 번호가 두 자리인가 (_01, _02)", ""),
    ("버전 표시 _v01 은 원본 파일에만 쓰이는가", ""),
    ("확장자가 알려주는 종류를 이름에 또 적지 않았는가", ""),
    ("오디오 카테고리 (Bgm/Sfx/Nar) 가 맨 앞에 _ 와 함께 있는가", ""),
    ("비디오 카테고리 (Info/Bgv) 가 맨 앞에 _ 와 함께 있는가", ""),
    ("자식 부품 FBX 에 부모 이름이 포함되어 있는가", ""),
    ("한글 파일명이 없는가 (영문만)", ""),
    ("애니메이션 포함 3D 모델에 접두사 Ani_ 가 붙어 있는가", ""),
]
r = write_table(ws, r, ["확인 항목", "체크 [✓]"], t1_check, col_widths=[68, 14])

r = section(ws, r, "[ T2 — Make Editor 인스턴스 ]")
t2_check = [
    ("타입 약자 (mdl_/img_/txt_/aud_ 등) 가 없는가", ""),
    ("모든 인스턴스명이 PascalCase 로 적혀 있는가", ""),
    ("이름 끝에 확장자 (.png/.fbx 등) 가 안 붙어 있는가", ""),
    ("구분자가 언더바 _ 하나만 쓰이는가 (이중 __ 없음, 하이픈 없음)", ""),
    ("상태가 정해진 목록 (_Pressed/_On/_Off/_Disabled · 기본 Default 생략) 안에 있는가", ""),
    ("방향이 정해진 단어 (_Left/_Right/_Up/_Down/_Clockwise/_AntiClockwise) 로 적혀 있는가", ""),
    ("방향 자식 노드(Left/Right/Up/Down)가 정해진 4방향 묶음에서만 쓰이는가", ""),
    ("UI 패밀리 약자 (Gnb/Snb/Bnb/Btn/Modal/Nav/Pnl/...) 를 사용했는가 (권장)", ""),
    ("숫자 경로(최대 3자리)가 하이픈 대신 언더바로 적혀 있는가 (Contents_3_1_1)", ""),
    ("텍스트 오브젝트가 직관적 이름이고, TTS 적용 시에만 접두사 TTS_ 가 붙어 있는가", ""),
    ("(v8) 메뉴 버튼을 메뉴명 그룹(Gnb_Home)으로 묶었는가 (번호는 이름 없을 때만)", ""),
    ("(v8) 컴포넌트 자식 이름이 부모_역할 형태인가 (_Img/_Text/_Ico/_Btn/_Bg, 'Text 1' 없음)", ""),
    ("(v8) 이벤트를 투명 *_Btn 에만 붙였고 여유 마진 + 맨 앞 Z(-0.02) 인가", ""),
    ("(v8) _Img(0)/_Text·_Ico(-0.01)/_Btn(-0.02) Z 로 겹침(잘림) 방지했는가", ""),
    ("(v8) 같은 바 메뉴에 정렬·간격 규칙(Direction/Align/Spacing)이 있는가", ""),
    ("(v8) 이름은 영문, 한글은 표시명(DisplayTitle)에 넣었는가", ""),
]
r = write_table(ws, r, ["확인 항목", "체크 [✓]"], t2_check, col_widths=[68, 14])

r = section(ws, r, "[ 애니메이션 ]")
anim_check = [
    ("애니메이션 이름이 <효과 이름>[_<방향>] 형태인가 (대상 오브젝트 이름 미포함)", ""),
    ("효과 이름이 정해진 목록 (Bounce/ZoomIn/Pop/Pulse/Spin/Slide/Flip/Fade/Highlight) 안에 있는가", ""),
    ("Color(색상) 효과를 3D 모델에 쓰지 않았는가 (UI·2D 전용)", ""),
    ("'new Clip' 같은 Unity 기본값 이름이 남아 있지 않은가", ""),
]
r = write_table(ws, r, ["확인 항목", "체크 [✓]"], anim_check, col_widths=[68, 14])

r = section(ws, r, "[ ID·경로·개인정보 ]")
priv_check = [
    ("모든 ResourceID 가 UUID v4 형식인가", ""),
    ("ResourceID 가 파일 전체에서 중복되지 않는가", ""),
    ("외부 공유본의 OriginPath 에 사용자 홈 경로 (C:\\Users\\, /home/) 가 없는가", ""),
    ("작성자 식별 정보 (이메일·계정명 등) 가 누설되지 않는가", ""),
]
r = write_table(ws, r, ["확인 항목", "체크 [✓]"], priv_check, col_widths=[68, 14])

# ─────────────────────────────────────────────────────────────
# Sheet 9 — Anti-Patterns
# ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Anti-Patterns")
r = title(ws, 1, "이렇게 쓰면 안 되는 예시")
r = note(ws, r, "실제 파일(Make Templete_20260512.make) 분석에서 찾은 잘못된 이름들입니다. v8 규칙에서는 모두 막거나 자동으로 고쳐집니다.", span=4)

anti = [
    ("노드 이름에 .png/.fbx 같은 확장자가 붙어 있음", "Logox4.png, btn_start.png", "Logo_x4, Btn_Start", "Make Editor 아이콘이 종류를 알려주므로 확장자 불필요"),
    ("공백이 들어간 노드 이름", "Text 1, 3D Model, Contents 1-1", "StartHeadline, Stage_3D, Contents_1_1", "공백을 언더바 _ 로 바꾸거나 의미 있는 이름으로"),
    ("이미지 이름이 알 수 없는 UUID 문자열", "6c34f19e-9d40-...", "Logo_x4 (의미 있는 이름)", "UUID 는 ID 로만 쓰고, 이름은 의미 있게"),
    ("'new Clip' 같은 Unity 기본값 애니메이션 이름", "new Clip, new Clip1, ...", "Spin_Left, Pop", "Unity 기본값 그대로 두지 말 것"),
    ("숫자로 시작하는 노드 이름", "0016", "Cupcake (용어집에 등록)", "이름은 영문 글자로 시작"),
    ("오타", "Btn_Strat", "Btn_Start", "오타 정정 (Strat → Start)"),
    ("하이픈 - 이 들어간 이름", "Snb_3-1-1, Contents 3-1-1", "Snb_3_1_1, Contents_3_1_1", "하이픈은 사용 금지"),
    ("OriginPath 에 작성자 PC 경로가 그대로 남음", "C:\\Users\\VIRNECT\\Desktop\\...", "<HOME>/... 또는 OriginPath 자체 제거", "개인정보·환경 정보 누설 방지"),
    ("폴더 경로에 한글이 들어감", "\\리소스\\모델\\...", "/resources/models/", "영문으로 통일, 한글 용어는 용어집에"),
    ("같은 이름이 형제 노드에 대량 중복", "형제 버튼 44개가 모두 'Btn'", "Gnb_01, Gnb_02 … (번호로 구별)", "형제 노드는 번호로 구별"),
]
fills_anti = [BAD_FILL] * len(anti)
r = write_table(ws, r, ["문제", "현재", "v7 에서는", "이유"], anti,
                col_widths=[44, 38, 38, 44])

# ─────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote: {OUT}")
print(f"Sheets: {wb.sheetnames}")
